import os
import openai
import openpyxl
from openpyxl.styles import Alignment
import time
import shutil
from datetime import datetime
import json
from tiktoken import encoding_for_model

# 读取配置文件
with open('config/config.json', 'r', encoding='utf-8') as config_file:
    config = json.load(config_file)

# 设置当前unit
file_path = __file__
current_unit = os.path.splitext(os.path.basename(file_path))[0]
# current_unit = "a1"

# 获取文件列表配置
files_config = config["files"]

# 查找当前unit的配置
current_file_config = next((file for file in files_config if os.path.splitext(file["input_file"])[0] == current_unit), None)

if current_file_config:
    input_file = current_file_config["input_file"]
    prompt_template1 = config["global"]["prompt_template"]
    system_message1 = config["global"]["system_message"]
else:
    raise ValueError(f"Configuration for {current_unit} not found.")

# 获取全局配置
global_config = config["global"]
root_path = global_config["root_path"]
apikey = global_config["apikey"]
apiurl = global_config["apiurl"]
model_name = global_config["model_name"]
input_col = global_config["input_col"]
output_col = global_config["output_col"]


#-----------------------------------------------------------------------------------------------#
# 定义路径
config_path = os.path.join(root_path, 'config', 'config.json')
input_dir = os.path.join(root_path, 'file')
output_dir = os.path.join(root_path, 'output')

#-----------------------------------------------------------------------------------------------#
# 设置OpenAI API密钥和URL
os.environ["OPENAI_API_KEY"] = apikey
os.environ["OPENAI_BASE_URL"] = apiurl

# 设置OpenAI API密钥
openai.api_key = os.getenv("OPENAI_API_KEY")

# 选择模型并获取编码器
model = model_name
encoder = encoding_for_model(model)

#-----------------------------------------------------------------------------------------------#

# 打印调试信息
print(f"当前脚本: {file_path}, 当前单元: {current_unit}, 输入文件: {input_file}")

xlsx_path = os.path.join(input_dir, input_file)
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# 添加标题
ws.cell(row=1, column=output_col, value="梳理结果")

#-----------------------------------------------------------------------------------------------#

def get_response(prompt_template, system_message, text):
    prompt = prompt_template + system_message + f"\n\n{text}\n\n" 
    max_retries = 3
    for attempt in range(max_retries):
        try:
            client = openai.OpenAI()
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "user", "content": prompt},
                ],
                max_tokens=300,
                n=1,
                temperature=0.5,
            )
            content = response.choices[0].message.content.strip()
            return content
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"Error occurred: {e}. Retrying...")
                time.sleep(20)  # 等待20秒后重试
            else:
                print(f"Failed after {max_retries} attempts: {e}")
                return "处理失败"


# 添加计数和进度显示
total_rows = ws.max_row - 1  # 除去标题行
processed_rows = 0
total_tokens = 0


#-----------------------------------------------------------------------------------------------#


# 从第{input_col}列读取数据并进行处理
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=input_col, max_col=input_col):
    for cell in row:
        if cell.value:
            text = cell.value
            # 计算当前文本的token数
            token_count = len(encoder.encode(text))
            total_tokens += token_count

            # 获取response
            response1 = get_response(prompt_template1, system_message1, text)
            print(f"判断1: {response1}")
            ws.cell(row=cell.row, column=output_col, value=response1)  # 将结果写入第{output_col}列

            processed_rows += 1

            # 显示进度
            progress_percentage = (processed_rows / total_rows) * 100
            print(f"正在处理第{input_file} 的第 {cell.row} 行，共{total_rows}行， 已完成: {progress_percentage:.2f}%")

            # 每处理一行，保存一次进度
            new_xlsx_path = os.path.join(output_dir, input_file)
            wb.save(new_xlsx_path)

#-----------------------------------------------------------------------------------------------#

# 最终保存到新的Excel文件
wb.save(new_xlsx_path)
print(f'API接口调用完成. 结果保存至 {new_xlsx_path}')

# 复制文件作为备份
current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
final_xlsx_path = os.path.join(output_dir, f'openAI_read_{current_unit}_{current_time}.xlsx')
shutil.copy(new_xlsx_path, final_xlsx_path)
print(f'副本 {final_xlsx_path} 已生成')

# 计算费用
price_per_1k_tokens = 0.002
total_cost = (total_tokens / 1000) * price_per_1k_tokens
print(f"累计Token: {total_tokens}")
print(f"累计花费: ${total_cost:.4f}")
